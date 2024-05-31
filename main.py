import json

import openpyxl

from concept import Concept
from concept_meta import ConceptName, ConceptDescription
from http_utils import post_concept, get_concept


def add_concepts_to_list(r):
    brand_proprietary_name = sheet.cell(row=r, column=4).value
    inn = sheet.cell(row=r, column=5).value
    strength = sheet.cell(row=r, column=6).value
    dosage_form = sheet.cell(row=r, column=7).value
    route_of_administration = sheet.cell(row=r, column=8).value
    pack_size = sheet.cell(row=r, column=9).value
    shelf_life = sheet.cell(row=r, column=13).value
    storage_conditions = sheet.cell(row=r, column=14).value
    registration_number = sheet.cell(row=r, column=16).value
    country_of_origin = sheet.cell(row=r, column=22).value
    atc_code = sheet.cell(row=r, column=23).value
    product_visual_descriptions = sheet.cell(row=r, column=30).value
    therapeutic_classification_level_3 = sheet.cell(row=r, column=24).value
    if therapeutic_classification_level_3:
        therapeutic_classification_level_3 = therapeutic_classification_level_3.title()
    if therapeutic_classification_level_3 and therapeutic_classification_level_3 not in sub_domains_list:
        sub_domains_list.append(therapeutic_classification_level_3)

    if brand_proprietary_name:
        pharm = Concept(brand_proprietary_name=brand_proprietary_name, inn=inn,
                        product_visual_descriptions=product_visual_descriptions)
        pharm.display_name = brand_proprietary_name
        pharm.extras = {"atc_code": atc_code, "strength": strength, "pack_size": pack_size,
                        "registration_number": registration_number}
        pharm.storage_conditions = storage_conditions
        pharm.dosage_form = dosage_form
        pharm.route_of_administration = route_of_administration
        pharm.shelf_life = shelf_life
        if country_of_origin:
            pharm.extras["country_of_origin"] = country_of_origin

        pharm.therapeutic_classification = therapeutic_classification_level_3
        pharm.pk = "{}-{}-{}-{}-{}-{}-{}-{}".format(brand_proprietary_name, inn, strength, dosage_form,
                                                    route_of_administration, shelf_life, storage_conditions,
                                                    registration_number)

        concepts_list.append(pharm)


def remove_duplicate_concepts():
    list_of_concepts_to_remove = []
    for concept_1 in concepts_list:
        for concept_2 in concepts_list:
            if (concept_1 != concept_2 and concept_1.pk == concept_2.pk
                    and concept_1.extras['pack_size'] != concept_2.extras['pack_size']):
                concept_1_pack_size = concept_1.extras['pack_size']
                concept_1.extras['pack_size'] = [concept_1_pack_size, concept_2.extras['pack_size']]
                concepts_list.remove(concept_2)


def upload_sub_domains():
    for tc in sub_domains_list:
        sub_domain = Concept(brand_proprietary_name=None, inn=None, product_visual_descriptions=None)
        sub_domain.concept_class = "Subdomain"
        sub_domain.datatype = "N/A"
        sub_domain.names = [ConceptName(name=tc, type="FULLY SPECIFIED")]
        sub_domain.descriptions = [ConceptDescription(tc)]
        sub_domain.type = "Concept"
        resp_dict = post_concept(concept=sub_domain, parent_id=None, parent_concept_name=None,
                                 parent_concept_url=None)
        sub_domains_dict[tc] = resp_dict


if __name__ == '__main__':
    wb = openpyxl.load_workbook("files/excel/PPB_Pharma_Product_31st_May_2024.xlsx")
    sheet = wb["Final_Work"]

    concepts_list = []
    sub_domains_list = []
    sub_domains_dict = {}
    for row in range(2, sheet.max_row + 1):
        add_concepts_to_list(row)

    remove_duplicate_concepts()

    upload_sub_domains()

    for concept in concepts_list:
        tcc = concept.therapeutic_classification
        tcc_details = sub_domains_dict[tcc]
        concept.parent_id = tcc_details['id']

        try:
            pass
            #concept.print_tree(0, "")
            post_concept(concept=concept, parent_id=tcc_details['id'], parent_concept_name=tcc_details['name'],
                         parent_concept_url=tcc_details['url'])
        except Exception as e:
            print("Exception ocurred - the error is: ", e)
            concept.print_tree(0, "")
            print("\n\n")
        finally:
            print("Done with pharm")
