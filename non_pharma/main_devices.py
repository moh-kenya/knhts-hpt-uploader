import openpyxl

from concept import Concept
from http_utils import post_concept, get_concept


if __name__ == '__main__':
    concept_id = 0
    wb = openpyxl.load_workbook("../files/excel/Final_MD_Pharma_Products_7th_June_4pm.xlsx")
    sheet = wb["MEDICAL_SUPPLIES"]

    medical_supplies = Concept("Medical Devices")
    medical_supplies.concept_class = "Subdomain"
    medical_supplies.datatype = "N/A"
    if concept_id:
        ms_dict = get_concept(concept_id)
    else:
        ms_dict = post_concept(
            concept=medical_supplies, parent_id=None,
            parent_concept_name=None, parent_concept_url=None)
    print(ms_dict)

    count = 1
    for row in range(2, sheet.max_row + 1):
        count += 1
        full_name = sheet.cell(row=row, column=5).value
        concept_class = sheet.cell(row=row, column=8).value
        # Extras
        registration_code = sheet.cell(row=row, column=2).value
        storage_condition = sheet.cell(row=row, column=19).value
        device_subset = sheet.cell(row=row, column=8).value
        shelf_life = sheet.cell(row=row, column=21).value
        country_of_origin = sheet.cell(row=row, column=25).value
        test_discipline = sheet.cell(row=row, column=33).value
        mah = sheet.cell(row=row, column=22).value
        manufacturer_name = sheet.cell(row=row, column=23).value
        application = sheet.cell(row=row, column=9).value
        # Brand
        trade_name = sheet.cell(row=row, column=4).value


        if full_name:
            non_pharm = Concept(full_name, trade_name)
            non_pharm.concept_class = concept_class
            non_pharm.extras = {"registration_code": registration_code,
                                "storage_condition": storage_condition,
                                "device_subset": device_subset,
                                "shelf_life": shelf_life, "MAH": mah,
                                "country_of_origin": country_of_origin,
                                "test_discipline": test_discipline,
                                "manufacturer_name": manufacturer_name,
                                "application": application}

            try:
                # non_pharm.print_tree(0, "")
                non_pharm.parent_id = ms_dict['id']
                pc_name = ms_dict['names'][0]['name']
                post_concept(concept=non_pharm, parent_id=ms_dict['id'],
                             parent_concept_name=pc_name,
                             parent_concept_url=ms_dict['url'])

                print(non_pharm)
            except Exception as e:
                print("The error is: ", e)
                non_pharm.print_tree(0, "")
                print("\n\n")
            finally:
                print("Done with non-pharm")
